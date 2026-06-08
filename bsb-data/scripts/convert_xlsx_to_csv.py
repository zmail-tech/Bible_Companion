#!/usr/bin/env python3
"""Convert BSB concordance XLSX to CSV format.

Usage:
    python3 scripts/convert_xlsx_to_csv.py sources/bsb_concordance/bsb_concordance.xlsx
"""

import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def convert_xlsx_to_csv(xlsx_path: str, csv_path: str | None = None) -> None:
    """Convert an XLSX file to CSV, matching the format expected by build_english_concordance.py."""
    xlsx = Path(xlsx_path)
    if not xlsx.exists():
        print(f"ERROR: File not found: {xlsx}")
        sys.exit(1)

    if csv_path is None:
        csv_path = str(xlsx.with_suffix(".csv"))

    print(f"Converting {xlsx} -> {csv_path}")

    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb[wb.sheetnames[0]]

    row_count = 0
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
            row_count += 1

    wb.close()
    print(f"Wrote {row_count:,} rows to {csv_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/convert_xlsx_to_csv.py <input.xlsx> [output.csv]")
        sys.exit(1)
    convert_xlsx_to_csv(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
